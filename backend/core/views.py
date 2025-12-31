# core/views.py
import csv
import io
from datetime import datetime, date
from django.db import transaction
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework import viewsets, permissions, status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import AllowAny
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import viewsets, filters, status
from .models import (Client, Customer, Rig, ServiceType,EquipmentType,Resource, Well,
                     HoleSection, Callout, SRO, Job, ExecutionLogEntry,Schedule,Asset,
                     EmployeeMaster,AssignedService,Field,CasingSize,DrillpipeSize,
                     MinimumIdSize,HoleSectionRelationship)
from .serializers import (
    ClientSerializer,
    CustomerSerializer,
    RigSerializer,
    CalloutSerializer,
    SroSerializer,
    JobSerializer,
    ExecutionLogEntrySerializer,
    RegisterSerializer,
    HoleSectionSerializer,
    ServiceTypeSerializer,
    WellSerializer,
    ScheduleSerializer,
    EquipmentTypeSerializer,
    ResourceSerializer,
    AssetSerializer,
   EmployeeSerializer,
   AssignedServiceListSerializer,
   AssignedServiceCreateSerializer,
   AssignedServiceUpdateSerializer,
   FieldSerializer,
   CasingSizeSerializer,
   DrillpipeSizeSerializer,
   MinimumIdSizeSerializer,
   CasingSizeSerializer,
   DrillpipeSizeSerializer,
   MinimumIdSizeSerializer,
   HoleSectionRelationshipSerializer




)
from openpyxl import load_workbook

def normalize_key(s: str) -> str:
    s = (
        str(s or "")
        .replace("\ufeff", "")
        .replace("\xa0", " ")
        .strip()
        .lower()
    )
    return " ".join(s.split()) 
def parse_date(value):
    """
    Supports inputs like:
    - 15/May/19
    - 02/Sep/22
    - 27/Apr/61
    - 2022-09-02
    - 02/09/2022
    """
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None

    formats = [
    "%d/%b/%y", "%d/%b/%Y",
    "%d-%b-%y", "%d-%b-%Y",   # ✅ add these two
    "%Y-%m-%d",
    "%d/%m/%Y", "%d/%m/%y",
    "%d-%m-%Y", "%d-%m-%y",
]
    for f in formats:
        try:
            return datetime.strptime(s, f).date()
        except ValueError:
            continue
    return None


def clean_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


# Map your column headers -> model fields.
# If your upload file has different header names, adjust here.
EMPLOYEE_HEADER_MAP = {
    "Emp #": "emp_number",
    "Name": "name",
    "Short Name": "short_name",
    "Designation": "designation",
    "Nationality": "nationality",
    "Date of Joining": "date_of_joining",
    "Grade": "grade",
    "Civil ID Number": "civil_id_number",
    "Civil ID Expiry Date": "civil_id_expiry_date",
    "Passport #": "passport_number",
    "Passport Expiry Date": "passport_expiry_date",
    "Visa Number": "visa_number",
    "Visa Issue Date": "visa_issue_date",
    "Visa Expiry Date": "visa_expiry_date",
    "Date of Birth": "date_of_birth",
    "Age": "age",
    "Blood Group": "blood_group",
    "Driving License": "driving_license",
    "DL Expiry Date": "dl_expiry_date",
    "Tel #": "tel_number",
    "Email ID": "email_id",
    "Address": "address",
    "Registered Bank Details": "registered_bank_details",
    "Acc Number": "acc_number",
    "Employee Type": "employee_type",
    "Contract Type": "contract_type",
    "Department": "department",
    "Leave Schedule": "leave_schedule",
    "Last Appraisal": "last_appraisal",
    "Benefits": "benefits",
    "Graduation": "graduation",
    "Specialization": "specialization",
    "Year of Passing": "year_of_passing",
    "University": "university",
    "Gender": "gender",
    "Marital Status": "marital_status",
    "Name of Spouse": "name_of_spouse",
    "No. of Kids": "number_of_kids",
    "Name Kid 1": "name_kid_1",
    "Name Kid 2": "name_kid_2",
    "Emergency Contact Name": "emergency_contact_name",
    "Emergency Contact Relationship": "emergency_contact_relationship",
    "Emergency Contact Tel #": "emergency_contact_tel",
    "Place of Birth": "place_of_birth",
}


DATE_FIELDS = {
    "date_of_joining",
    "civil_id_expiry_date",
    "passport_expiry_date",
    "visa_issue_date",
    "visa_expiry_date",
    "date_of_birth",
    "dl_expiry_date",
    "last_appraisal",
}


INT_FIELDS = {"age", "year_of_passing", "number_of_kids"}



# =============================ASSET======================================#
def _norm_header(s: str) -> str:
    return (s or "").strip().lower().replace("\ufeff", "")

def _parse_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(str(v).strip(), fmt).date()
        except Exception:
            pass
    return None

def _parse_float(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except Exception:
        return None

def _status_normalize(v: str) -> str:
    s = (v or "").strip().lower()
    mapping = {
        "on duty": "on_duty",
        "onduty": "on_duty",
        "on_duty": "on_duty",
        "yellow": "yellow",
        "green": "green",
        "upgraded": "upgraded",
        "off duty": "off_duty",
        "off_duty": "off_duty",
        "maintenance": "maintenance",
        "breakdown": "breakdown",
    }
    return mapping.get(s, s.replace(" ", "_"))

HEADER_MAP = {
    "account code": "account_code",
    "asset code": "asset_code",
    "status": "status",
    "cost center": "cost_center",
    "department": "department",
    "asset group": "asset_group",
    "physical location": "physical_location",
    "asset main category": "asset_main_category",
    "asset sub category": "asset_sub_category",
    "asset description": "asset_description",
    "serial no.": "serial_no",
    "serial no": "serial_no",
    "part #": "part_no",
    "part#": "part_no",
    "mfg serial": "mfg_serial",
    "manufacturer": "manufacturer",
    "comments": "comments",
    "certificate": "certificate",
    "last regular inspection date": "last_regular_inspection_date",
    "next due date for regular inspection": "next_due_date_regular_inspection",
    "last minor service date": "last_minor_service_date",
    "next due date for minor service": "next_due_date_minor_service",
    "last major service date": "last_major_service_date",
    "next due date for major service": "next_due_date_major_service",
    "last regular inspection km": "last_regular_inspection_km",
    "last regular inspection hours": "last_regular_inspection_hours",
    "last regular inspection jobs": "last_regular_inspection_jobs",
    "last minor service  km": "last_minor_service_km",
    "last minor service km": "last_minor_service_km",
    "last minor service hours": "last_minor_service_hours",
    "last minor service jobs": "last_minor_service_jobs",
    "last major service km": "last_major_service_km",
    "last major service hours": "last_major_service_hours",
    "last major service jobs": "last_major_service_jobs",
    "next regular inspection km": "next_regular_inspection_km",
    "next regular inspection jobs": "next_regular_inspection_jobs",
    "next regular inspection hour": "next_regular_inspection_hour",
    "next minor service km": "next_minor_service_km",
    "next minor service jobs": "next_minor_service_jobs",
    "next minor service hours": "next_minor_service_hours",
    "next major service jobs": "next_major_service_jobs",
    "next major service km": "next_major_service_km",
    "next major service hours": "next_major_service_hours",
    "third-party comment": "third_party_comment",
    "third-party service date": "third_party_service_date",
    "attachments1": "attachments1",
    "breakdown maintenance comment": "breakdown_maintenance_comment",
    "next third-party service date": "next_third_party_service_date",
    "maintenance type": "maintenance_type",
    "ncr- no": "ncr_no",
    "ncr- no.": "ncr_no",
    "ncr - no": "ncr_no",
}

# =============================ASSET======================================#

class RegisterView(APIView):
    permission_classes = [AllowAny]

    def post(self, request, *args, **kwargs):
        serializer = RegisterSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.save()

            # OPTIONAL: automatically issue JWT tokens on registration
            refresh = RefreshToken.for_user(user)
            return Response(
                {
                    "user": {
                        "id": user.id,
                        "username": user.username,
                        "email": user.email,
                        "first_name": user.first_name,
                        "last_name": user.last_name,
                    },
                    "access": str(refresh.access_token),
                    "refresh": str(refresh),
                },
                status=status.HTTP_201_CREATED,
            )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# Base viewset with authentication
class BaseViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]

# All your viewset classes
class ClientViewSet(BaseViewSet):
    queryset = Client.objects.all()
    serializer_class = ClientSerializer

class CustomerViewSet(BaseViewSet):  # Added BaseViewSet inheritance
    queryset = Customer.objects.all()
    serializer_class = CustomerSerializer

class RigViewSet(BaseViewSet):
    queryset = Rig.objects.all()
    serializer_class = RigSerializer

class ResourceViewSet(BaseViewSet):
    queryset = Resource.objects.all()
    serializer_class = ResourceSerializer

class EquipmentTypeViewSet(BaseViewSet):
    queryset = EquipmentType.objects.all()
    serializer_class = EquipmentTypeSerializer

class FieldViewSet(BaseViewSet):
    queryset = Field.objects.all()
    serializer_class =  FieldSerializer

    
class WellViewSet(BaseViewSet):  # Added BaseViewSet inheritance
    queryset = Well.objects.all()
    serializer_class = WellSerializer

class ServiceTypeViewSet(BaseViewSet):  # Added BaseViewSet inheritance
    queryset = ServiceType.objects.all()
    serializer_class = ServiceTypeSerializer

# core/views.py - Update the get_available_options method

class HoleSectionViewSet(BaseViewSet):
    queryset = HoleSection.objects.all()
    serializer_class = HoleSectionSerializer
    
    @action(detail=True, methods=['get'], url_path='available-options')
    def get_available_options(self, request, pk=None):
        """Get available pipe sizes for a specific hole section"""
        hole_section = self.get_object()
        
        # Parse hole section size from name (e.g., "16" from "16 inch" or "12 1/4")
        hole_section_size = self._parse_hole_section_size(hole_section.name)
        
        # Initialize data
        data = {
            'available_casing_sizes': [],
            'available_drillpipe_sizes': [],
            'available_minimum_ids': [],
        }
        
        if hole_section_size:
            # Get relationships for this hole section if they exist
            relationships = hole_section.relationships.all()
            
            if relationships.exists():
                # Use defined relationships
                relationship = relationships.first()
                data['available_casing_sizes'] = CasingSizeSerializer(
                    relationship.allowed_casing_sizes.all(), many=True
                ).data
                data['available_drillpipe_sizes'] = DrillpipeSizeSerializer(
                    relationship.allowed_drillpipe_sizes.all(), many=True
                ).data
            else:
                # If no relationships defined, filter by size comparison
                # Get all casing sizes SMALLER than hole section
                smaller_casing_sizes = CasingSize.objects.filter(
                    size__lt=hole_section_size
                ).order_by('-size')  # Order by largest first
                
                # Get all drillpipe sizes SMALLER than hole section
                smaller_drillpipe_sizes = DrillpipeSize.objects.filter(
                    size__lt=hole_section_size
                ).order_by('-size')
                
                data['available_casing_sizes'] = CasingSizeSerializer(
                    smaller_casing_sizes, many=True
                ).data
                data['available_drillpipe_sizes'] = DrillpipeSizeSerializer(
                    smaller_drillpipe_sizes, many=True
                ).data
            
            # Get minimum ID sizes that are smaller than the smallest available pipe
            # First get the smallest available pipe size
            smallest_pipe_size = None
            if data['available_casing_sizes']:
                casing_sizes = [cs['size'] for cs in data['available_casing_sizes']]
                smallest_pipe_size = min(casing_sizes) if casing_sizes else None
            
            if data['available_drillpipe_sizes']:
                drillpipe_sizes = [dp['size'] for dp in data['available_drillpipe_sizes']]
                drillpipe_min = min(drillpipe_sizes) if drillpipe_sizes else None
                if smallest_pipe_size is None or (drillpipe_min and drillpipe_min < smallest_pipe_size):
                    smallest_pipe_size = drillpipe_min
            
            if smallest_pipe_size:
                # Get minimum ID sizes smaller than the smallest pipe
                min_id_sizes = MinimumIdSize.objects.filter(
                    size__lt=smallest_pipe_size
                ).order_by('-size')
            else:
                # If no pipe sizes available, return all minimum IDs
                min_id_sizes = MinimumIdSize.objects.all()
            
            data['available_minimum_ids'] = MinimumIdSizeSerializer(
                min_id_sizes, many=True
            ).data
        else:
            # If we can't parse hole section size, return all sizes
            data['available_casing_sizes'] = CasingSizeSerializer(
                CasingSize.objects.all(), many=True
            ).data
            data['available_drillpipe_sizes'] = DrillpipeSizeSerializer(
                DrillpipeSize.objects.all(), many=True
            ).data
            data['available_minimum_ids'] = MinimumIdSizeSerializer(
                MinimumIdSize.objects.all(), many=True
            ).data
        
        return Response(data)
    
    def _parse_hole_section_size(self, name: str):
        """Parse numeric size from hole section name"""
        import re
        from decimal import Decimal
        
        if not name:
            return None
            
        # Try to extract numbers (integers and decimals)
        # This handles formats like: "16", "12 1/4", "8.5", "9 5/8"
        match = re.search(r'(\d+(?:\.\d+)?)(?:\s+(\d+)/(\d+))?', str(name))
        
        if not match:
            return None
            
        whole_part = match.group(1)
        numerator = match.group(2)
        denominator = match.group(3)
        
        try:
            size = Decimal(whole_part)
            
            # Add fraction if present (e.g., "12 1/4" -> 12.25)
            if numerator and denominator:
                fraction = Decimal(numerator) / Decimal(denominator)
                size += fraction
                
            return size
        except:
            return None

class CasingSizeViewSet(viewsets.ModelViewSet):
    queryset = CasingSize.objects.all()
    serializer_class = CasingSizeSerializer
    permission_classes = [IsAuthenticated]

class DrillpipeSizeViewSet(viewsets.ModelViewSet):
    queryset = DrillpipeSize.objects.all()
    serializer_class = DrillpipeSizeSerializer
    permission_classes = [IsAuthenticated]

class MinimumIdSizeViewSet(viewsets.ModelViewSet):
    queryset = MinimumIdSize.objects.all()
    serializer_class = MinimumIdSizeSerializer
    permission_classes = [IsAuthenticated]

class HoleSectionRelationshipViewSet(viewsets.ModelViewSet):
    queryset = HoleSectionRelationship.objects.all()
    serializer_class = HoleSectionRelationshipSerializer
    permission_classes = [IsAuthenticated]
    
    @action(detail=True, methods=['get'], url_path='available-options')
    def get_available_options(self, request, pk=None):
        hole_section = self.get_object()
        
        # Get relationships for this hole section
        relationships = hole_section.relationships.all()
        
        data = {
            'available_casing_sizes': [],
            'available_drillpipe_sizes': [],
        }
        
        if relationships.exists():
            relationship = relationships.first()
            data['available_casing_sizes'] = CasingSizeSerializer(
                relationship.allowed_casing_sizes.all(), many=True
            ).data
            data['available_drillpipe_sizes'] = DrillpipeSizeSerializer(
                relationship.allowed_drillpipe_sizes.all(), many=True
            ).data
        
        return Response(data)

class CalloutViewSet(BaseViewSet):
    queryset = Callout.objects.all()
    serializer_class = CalloutSerializer
    permission_classes = [IsAuthenticated] 
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=["post"], url_path="generate-sro")
    def generate_sro(self, request, pk=None):
        """
        POST /callouts/{id}/generate-sro/
        """
        callout = self.get_object()

        # Already has SRO?
        if hasattr(callout, "sro"):
            return Response(
                {"detail": "SRO already exists for this callout."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Optional: add some validations here (e.g. callout must be locked)

        sro = SRO.objects.create(
            callout=callout,
            created_by=request.user if request.user.is_authenticated else None,
        )

        # update callout status → SRO Activated
        callout.status = "sro_activated"
        callout.save(update_fields=["status"])

        serializer = SroSerializer(sro)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class SroViewSet(BaseViewSet):
    queryset = SRO.objects.all().select_related("callout")
    serializer_class = SroSerializer

    @action(detail=True, methods=["post"], url_path="approve")
    def approve(self, request, pk=None):
        """
        POST /sros/{id}/approve/

        ✅ Only APPROVES the SRO (no schedule creation here).
        Frontend will then show "Schedule Service".
        Schedule creation will later switch SRO -> scheduled.
        """
        sro = self.get_object()

        # If already scheduled, keep it as is
        if sro.status == "scheduled":
            return Response(self.get_serializer(sro).data, status=status.HTTP_200_OK)

        # If already approved, keep it as is
        if sro.status == "approved":
            return Response(self.get_serializer(sro).data, status=status.HTTP_200_OK)

        # ✅ Approve here (NOT scheduled)
        sro.status = "approved"
        sro.save(update_fields=["status"])

        # ✅ Optional: update callout status when SRO approved
        # (only if you want this behavior)
        if sro.callout and sro.callout.status != "sro_activated":
            sro.callout.status = "sro_activated"
            sro.callout.save(update_fields=["status"])

        return Response(self.get_serializer(sro).data, status=status.HTTP_200_OK)
class ScheduleViewSet(BaseViewSet):
    queryset = Schedule.objects.select_related("sro", "sro__callout").all()
    serializer_class = ScheduleSerializer

    
    def perform_create(self, serializer):
        schedule = serializer.save(created_by=self.request.user)


        # ✅ When schedule is created -> mark Callout as scheduled
        callout = schedule.sro.callout
        callout.status = "scheduled"
        callout.save(update_fields=["status"])



class JobViewSet(BaseViewSet):
    queryset = Job.objects.select_related("sro").all()
    serializer_class = JobSerializer

class ExecutionLogEntryViewSet(BaseViewSet):
    queryset = ExecutionLogEntry.objects.select_related("job").all()
    serializer_class = ExecutionLogEntrySerializer

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)



class AssetViewSet(BaseViewSet):
    queryset = Asset.objects.all().order_by("asset_code")
    serializer_class = AssetSerializer

    @action(detail=False, methods=["get"], url_path="by-cost-centers")
    def by_cost_centers(self, request):
        raw = request.query_params.get("cost_centers", "")
        cost_centers = [x.strip() for x in raw.split(",") if x.strip()]
        qs = Asset.objects.filter(cost_center__in=cost_centers).order_by("asset_code")
        data = self.get_serializer(qs, many=True).data
        return Response(data)

    @action(detail=False, methods=["get"], url_path="green-cost-centers")
    def green_cost_centers(self, request):
        cost_centers = (
            Asset.objects.filter(status="green")
            .exclude(cost_center__isnull=True)
            .exclude(cost_center__exact="")
            .values_list("cost_center", flat=True)
            .distinct()
            .order_by("cost_center")
        )
        return Response(list(cost_centers))

    @action(
        detail=False,
        methods=["post"],
        url_path="import",
        parser_classes=[MultiPartParser, FormParser],
    )
    def import_assets(self, request):
        uploaded = request.FILES.get("file")
        if not uploaded:
            return Response({"detail": "file is required"}, status=status.HTTP_400_BAD_REQUEST)

        name = (uploaded.name or "").lower()

        try:
            if name.endswith(".csv"):
                rows = self._read_csv(uploaded)
            elif name.endswith(".xlsx"):
                rows = self._read_xlsx(uploaded)
            else:
                return Response(
                    {"detail": "Only .csv or .xlsx files are supported"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except Exception as e:
            return Response(
                {"detail": f"Failed to read file: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created_count = 0
        updated_count = 0
        unchanged_count = 0
        errors = []
        preview = []

        def should_skip(v):
            # Skip None / empty string so we DON'T wipe existing values
            if v is None:
                return True
            if isinstance(v, str) and v.strip() == "":
                return True
            return False

        with transaction.atomic():
            for idx, raw in enumerate(rows, start=2):
                asset_code = (raw.get("asset_code") or "").strip()
                if not asset_code:
                    errors.append({"row": idx, "error": "Asset Code is required"})
                    continue

                if "status" in raw:
                    raw["status"] = _status_normalize(raw.get("status"))

                for k in [
                    "last_regular_inspection_date",
                    "next_due_date_regular_inspection",
                    "last_minor_service_date",
                    "next_due_date_minor_service",
                    "last_major_service_date",
                    "next_due_date_major_service",
                    "third_party_service_date",
                    "next_third_party_service_date",
                ]:
                    if k in raw:
                        raw[k] = _parse_date(raw.get(k))

                for k in [
                    "last_regular_inspection_km",
                    "last_regular_inspection_hours",
                    "last_regular_inspection_jobs",
                    "last_minor_service_km",
                    "last_minor_service_hours",
                    "last_minor_service_jobs",
                    "last_major_service_km",
                    "last_major_service_hours",
                    "last_major_service_jobs",
                    "next_regular_inspection_km",
                    "next_regular_inspection_jobs",
                    "next_regular_inspection_hour",
                    "next_minor_service_km",
                    "next_minor_service_jobs",
                    "next_minor_service_hours",
                    "next_major_service_jobs",
                    "next_major_service_km",
                    "next_major_service_hours",
                ]:
                    if k in raw:
                        raw[k] = _parse_float(raw.get(k))

                try:
                    obj = Asset.objects.filter(asset_code=asset_code).first()

                    if obj:
                        changed_fields = []

                        for field, value in raw.items():
                            if field == "asset_code":
                                continue
                            if should_skip(value):
                                continue

                            old = getattr(obj, field, None)

                            # compare strings safely
                            if isinstance(old, str) and isinstance(value, str):
                                old_cmp = old.strip()
                                new_cmp = value.strip()
                            else:
                                old_cmp = old
                                new_cmp = value

                            if old_cmp != new_cmp:
                                setattr(obj, field, value)
                                changed_fields.append(field)

                        if changed_fields:
                            obj.save(update_fields=changed_fields)
                            updated_count += 1
                            action_type = "updated"
                        else:
                            unchanged_count += 1
                            action_type = "unchanged"

                    else:
                        # Optional: you can also remove blank keys here before create if you want
                        obj = Asset.objects.create(**raw)
                        created_count += 1
                        action_type = "created"

                    preview.append(
                        {
                            "row": idx,
                            "action": action_type,
                            "asset_code": obj.asset_code,
                            "cost_center": getattr(obj, "cost_center", None),
                            "department": getattr(obj, "department", None),
                            "asset_group": getattr(obj, "asset_group", None),
                            "status": getattr(obj, "status", None),
                        }
                    )

                except Exception as e:
                    errors.append({"row": idx, "asset_code": asset_code, "error": str(e)})

        # ✅ THIS RETURN WAS MISSING / BROKEN IN YOUR VERSION
        return Response(
            {
                "created": created_count,
                "updated": updated_count,
                "unchanged": unchanged_count,
                "errors": errors[:200],
                "total_rows": len(rows),
                "preview": preview[:200],
            },
            status=status.HTTP_200_OK,
        )


    def _read_csv(self, uploaded_file):
        content = uploaded_file.read().decode("utf-8", errors="ignore")
        f = io.StringIO(content)
        reader = csv.DictReader(f)
        out = []
        for row in reader:
            mapped = {}
            for h, v in row.items():
                key = HEADER_MAP.get(_norm_header(h))
                if key:
                    mapped[key] = (v.strip() if isinstance(v, str) else v)
            out.append(mapped)
        return out

    def _read_xlsx(self, uploaded_file):
        wb = load_workbook(uploaded_file, data_only=True)
        ws = wb.active

        headers = []
        for cell in ws[1]:
            headers.append(_norm_header(cell.value if cell.value is not None else ""))

        out = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            mapped = {}
            for i, val in enumerate(r):
                excel_h = headers[i] if i < len(headers) else ""
                key = HEADER_MAP.get(excel_h)
                if key:
                    mapped[key] = val
            out.append(mapped)
        return out

class EmployeeViewSet(viewsets.ModelViewSet):
    queryset = EmployeeMaster.objects.all().order_by("emp_number")
    serializer_class = EmployeeSerializer


class EmployeeBulkUploadAPIView(APIView):
    parser_classes = [MultiPartParser, FormParser]

    @transaction.atomic
    def post(self, request):
        upload = request.FILES.get("file")
        if not upload:
            return Response({"detail": "file is required."}, status=status.HTTP_400_BAD_REQUEST)

        name = (upload.name or "").lower()

        # ---------------- XLSX ----------------
        if name.endswith(".xlsx"):
            wb = load_workbook(upload, read_only=True, data_only=True)
            ws = wb.active

            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                return Response({"detail": "Empty sheet."}, status=status.HTTP_400_BAD_REQUEST)

            # ✅ find header row (must contain "Emp #" and "Name")
            header_idx = None
            for i in range(min(30, len(all_rows))):
                row = all_rows[i] or []
                cells = [normalize_key(c) for c in row]
                if ("emp #" in cells or "emp#" in cells) and "name" in cells:
                    header_idx = i
                    break

            if header_idx is None:
                return Response(
                    {
                        "detail": "Header row not found. Make sure sheet contains 'Emp #' and 'Name'.",
                        "first_5_rows": [
                            [str(x) if x is not None else "" for x in (all_rows[j] or [])]
                            for j in range(min(5, len(all_rows)))
                        ],
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            headers = [str(h).strip() if h is not None else "" for h in all_rows[header_idx]]
            data_rows = all_rows[header_idx + 1 :]

            parsed = []
            for r in data_rows:
                # skip empty rows
                if not r or not any(v is not None and str(v).strip() for v in r):
                    continue
                row_dict = {headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))}
                parsed.append(row_dict)

            return self._process_rows(parsed, start_row=header_idx + 2)

        # ---------------- CSV / TSV ----------------
        if name.endswith(".csv") or name.endswith(".tsv"):
            raw = upload.read().decode("utf-8-sig", errors="ignore")
            sio = io.StringIO(raw)

            # detect delimiter
            delimiter = "\t" if name.endswith(".tsv") else ","
            try:
                dialect = csv.Sniffer().sniff(raw[:5000], delimiters=[",", "\t", ";", "|"])
                delimiter = dialect.delimiter
            except Exception:
                pass

            reader = csv.DictReader(sio, delimiter=delimiter)

            if reader.fieldnames:
                reader.fieldnames = [f.strip() if f else "" for f in reader.fieldnames]

            rows = list(reader)

            # handle "one big column" case
            if rows and len(rows[0].keys()) == 1:
                only_key = list(rows[0].keys())[0]
                sep = "\t" if "\t" in only_key else ";" if ";" in only_key else "," if "," in only_key else None
                if sep:
                    rows = self._split_single_column_rows(rows, sep=sep)

            return self._process_rows(rows, start_row=2)

        return Response(
            {"detail": "Unsupported file type. Upload .csv/.tsv or .xlsx only."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    def _split_single_column_rows(self, rows, sep="\t"):
        fixed = []
        header_key = list(rows[0].keys())[0]
        headers = [h.strip() for h in str(header_key).split(sep)]
        for r in rows:
            big_val = r.get(header_key) or ""
            values = [v.strip() for v in str(big_val).split(sep)]
            fixed.append({headers[i]: (values[i] if i < len(values) else None) for i in range(len(headers))})
        return fixed

    def _process_rows(self, rows, start_row=2):
        created = 0
        updated = 0
        errors = []

        for i, row in enumerate(rows):
            excel_row_number = start_row + i
            try:
                if not row or not any(v is not None and str(v).strip() for v in row.values()):
                    continue

                payload = self._row_to_payload(row)

                emp_number = payload.get("emp_number")
                if not emp_number:
                    raise ValueError("Emp # is missing.")

                # ✅ Name required
                if not payload.get("name"):
                    raise ValueError("Name is missing.")

                obj = EmployeeMaster.objects.filter(emp_number=emp_number).first()
                if obj:
                    ser = EmployeeSerializer(obj, data=payload, partial=True)
                    ser.is_valid(raise_exception=True)
                    ser.save()
                    updated += 1
                else:
                    ser = EmployeeSerializer(data=payload)
                    ser.is_valid(raise_exception=True)
                    ser.save()
                    created += 1

            except Exception as e:
                errors.append({"row": excel_row_number, "error": str(e)})

        return Response(
            {
                "total_rows": len(rows),
                "created": created,
                "updated": updated,
                "failed": len(errors),
                "errors": errors[:50],
            },
            status=status.HTTP_200_OK if not errors else status.HTTP_207_MULTI_STATUS,
        )

    def _row_to_payload(self, row):
        # normalize row headers
        nrow = {normalize_key(k): row.get(k) for k in row.keys()}

        # normalize template headers once
        norm_map = {normalize_key(h): field for h, field in EMPLOYEE_HEADER_MAP.items()}

        payload = {}
        for nheader, field in norm_map.items():
            val = nrow.get(nheader)

            if field in DATE_FIELDS:
                payload[field] = parse_date(val)
            elif field in INT_FIELDS:
                s = clean_str(val)
                payload[field] = int(s) if s and s.isdigit() else None
            else:
                payload[field] = clean_str(val)

            # ✅ fix scientific notation for acc_number from XLSX
            if field == "acc_number" and val is not None:
                try:
                    if isinstance(val, (int, float)):
                        payload[field] = str(int(val))
                    elif isinstance(val, str) and ("e" in val.lower()):
                        payload[field] = str(int(float(val)))
                except Exception:
                    pass

        if payload.get("place_of_birth"):
            payload["place_of_birth"] = payload["place_of_birth"].replace("#", "").strip()

        return payload

class AssignedServiceViewSet(viewsets.ModelViewSet):
    queryset = AssignedService.objects.all().order_by("-created_at")
    permission_classes = [IsAuthenticated]
    def perform_create(self, serializer):
        assigned_service = serializer.save()
        schedule = assigned_service.schedule
        if not schedule:
            return

        if schedule.status != "assigned":
            schedule.status = "assigned"
            schedule.save(update_fields=["status"])

        sro = getattr(schedule, "sro", None)
        if sro and sro.status != "assigned":
            sro.status = "assigned"
            sro.save(update_fields=["status"])

        callout = getattr(sro, "callout", None) if sro else None
        if callout and callout.status != "assigned":
            callout.status = "assigned"
        callout.save(update_fields=["status"])
    @action(detail=False, methods=["get"], url_path="by-schedule/(?P<schedule_id>[^/.]+)")
    def by_schedule(self, request, schedule_id=None):
        obj = AssignedService.objects.filter(schedule_id=schedule_id).first()
        if not obj:
            return Response({"detail": "No assigned service for this schedule."}, status=status.HTTP_404_NOT_FOUND)
        # return the normal serializer you use for retrieve/list
        ser = AssignedServiceListSerializer(obj)  # or a detail serializer if you have one
        return Response(ser.data)

    @action(detail=False, methods=["get"], url_path="approved-schedules")
    def approved_schedules(self, request):
        qs = Schedule.objects.filter(status="approved").filter(assigned_service__isnull=True)
        return Response(ScheduleSerializer(qs, many=True).data)

    def get_serializer_class(self):
        if self.action == "create":
            return AssignedServiceCreateSerializer

        # ✅ THIS FIXES EDIT
        if self.action in ["update", "partial_update"]:
            return AssignedServiceUpdateSerializer

        # retrieve/list can use list serializer (or create a detail serializer)
        return AssignedServiceListSerializer